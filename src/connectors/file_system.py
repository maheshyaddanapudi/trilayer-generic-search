from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator
from uuid import uuid4

from src.domain.connector import SourceConnector
from src.models.core import RawEntity, RawRelationship

log = logging.getLogger(__name__)

_PDF_MIME = "application/pdf"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SUPPORTED_MIMES = {_PDF_MIME, _DOCX_MIME, _XLSX_MIME}


@dataclass
class RawSection:
    heading: str
    content_text: str
    page_number: int
    order_index: int

    @property
    def content_summary(self) -> str:
        return self.content_text[:300].strip()


class TextExtractor:
    def extract_pdf(self, data: bytes) -> list[RawSection]:
        from pdfminer.high_level import extract_pages
        from pdfminer.layout import LTTextBox, LTTextLine
        import re

        sections: list[RawSection] = []
        current_heading = ""
        current_lines: list[str] = []
        page_num = 0

        for page_layout in extract_pages(io.BytesIO(data)):
            page_num += 1
            for element in page_layout:
                if not isinstance(element, LTTextBox):
                    continue
                for line in element:
                    if not isinstance(line, LTTextLine):
                        continue
                    text = line.get_text().strip()
                    if not text:
                        continue
                    if re.match(r'^§\d', text) or (text.isupper() and len(text.split()) <= 8):
                        if current_heading or current_lines:
                            sections.append(RawSection(
                                heading=current_heading or "Untitled",
                                content_text=" ".join(current_lines),
                                page_number=page_num,
                                order_index=len(sections),
                            ))
                        current_heading = text
                        current_lines = []
                    else:
                        current_lines.append(text)

        if current_heading or current_lines:
            sections.append(RawSection(
                heading=current_heading or "Untitled",
                content_text=" ".join(current_lines),
                page_number=page_num,
                order_index=len(sections),
            ))
        return sections

    def extract_docx(self, data: bytes) -> list[RawSection]:
        from docx import Document

        doc = Document(io.BytesIO(data))
        sections: list[RawSection] = []
        current_heading = ""
        current_paras: list[str] = []
        heading_styles = {"Heading 1", "Heading 2", "Heading 3"}

        for para in doc.paragraphs:
            if para.style.name in heading_styles:
                if current_heading or current_paras:
                    sections.append(RawSection(
                        heading=current_heading or "Untitled",
                        content_text=" ".join(current_paras),
                        page_number=1,
                        order_index=len(sections),
                    ))
                current_heading = para.text.strip()
                current_paras = []
            else:
                text = para.text.strip()
                if text:
                    current_paras.append(text)

        if current_heading or current_paras:
            sections.append(RawSection(
                heading=current_heading or "Untitled",
                content_text=" ".join(current_paras),
                page_number=1,
                order_index=len(sections),
            ))
        return sections

    def extract_xlsx(self, data: bytes) -> list[RawSection]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sections: list[RawSection] = []
        for idx, ws in enumerate(wb.worksheets):
            rows: list[str] = []
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                row_text = "\t".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
            sections.append(RawSection(
                heading=ws.title,
                content_text="\n".join(rows),
                page_number=idx + 1,
                order_index=idx,
            ))
        return sections

    def _detect_sections_heuristic(self, lines: list[str]) -> list[RawSection]:
        return []


class FileSystemConnector(SourceConnector):
    supports_incremental = False

    def __init__(self, watch_dir: Path, supported_mimes: list[str] | None = None,
                 domain_id: str = "documents") -> None:
        self._watch_dir = watch_dir
        self._supported_mimes = supported_mimes or list(SUPPORTED_MIMES)
        self._domain_id = domain_id
        self._extractor = TextExtractor()

    def connect(self) -> None:
        self._watch_dir.mkdir(parents=True, exist_ok=True)

    def disconnect(self) -> None:
        pass

    def fetch_all(self) -> Iterator[RawEntity]:
        for path in self._watch_dir.iterdir():
            if path.is_file():
                mime = self._mime_for(path)
                if mime:
                    yield from self.process_upload(path.read_bytes(), path.name, mime)

    def process_upload(self, file_bytes: bytes, filename: str,
                       mime_type: str) -> Iterator[RawEntity]:
        if mime_type not in self._supported_mimes:
            raise ValueError(f"Unsupported MIME type: {mime_type}")

        doc_id = str(uuid4())
        title = Path(filename).stem
        sections = self._extract_sections(file_bytes, mime_type, title)

        yield self._make_document_entity(doc_id, title, mime_type, len(sections))

        for section in sections:
            yield self._make_section_entity(section, doc_id, title)

    def _extract_sections(self, data: bytes, mime: str, title: str) -> list[RawSection]:
        if mime == _PDF_MIME:
            sections = self._extractor.extract_pdf(data)
        elif mime == _DOCX_MIME:
            sections = self._extractor.extract_docx(data)
        elif mime == _XLSX_MIME:
            sections = self._extractor.extract_xlsx(data)
        else:
            sections = []

        if not sections:
            sections = [RawSection(heading=title, content_text="", page_number=1, order_index=0)]
        return sections

    def _make_document_entity(self, doc_id: str, title: str,
                               mime_type: str, section_count: int) -> RawEntity:
        return RawEntity(
            entity_id=doc_id, entity_type="Document",
            domain_id=self._domain_id, name=title,
            description=f"Uploaded document: {title}",
            properties={
                "doc_id": doc_id, "title": title,
                "mime_type": mime_type, "section_count": section_count,
            },
        )

    def _make_section_entity(self, section: RawSection, doc_id: str,
                              doc_title: str) -> RawEntity:
        section_id = str(uuid4())
        heading = f"{doc_title} §{section.order_index + 1}" if section.heading == "Untitled" else section.heading
        return RawEntity(
            entity_id=section_id, entity_type="Section",
            domain_id=self._domain_id,
            name=heading,
            description=section.content_summary,
            properties={
                "section_id": section_id, "doc_id": doc_id,
                "heading": heading, "order_index": section.order_index,
                "page_number": section.page_number,
                "content_summary": section.content_summary,
            },
            parent_id=doc_id,
            relationships=[RawRelationship(source_id=doc_id, target_id=section_id,
                                           relation_type="HAS_SECTION")],
        )

    @staticmethod
    def _mime_for(path: Path) -> str | None:
        suffix = path.suffix.lower()
        return {
            ".pdf": _PDF_MIME,
            ".docx": _DOCX_MIME,
            ".xlsx": _XLSX_MIME,
        }.get(suffix)
